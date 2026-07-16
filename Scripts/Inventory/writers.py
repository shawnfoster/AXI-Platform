"""
AXI Inventory Engine - Writers
"""
from pathlib import Path
import csv
from openpyxl import load_workbook

HEADERS=["Inventory ID","Filename","Extension","Relative Path","Repository","File Size","Last Modified","Source Package","Object Type","Version","Checksum","Duplicate Candidate","Review Required","Lifecycle","Status","Notes"]

def write_csv(records,outfile:Path):
    with outfile.open("w",newline="",encoding="utf-8") as f:
        w=csv.writer(f)
        w.writerow(HEADERS)
        for r in records:
            d=r.to_dict()
            w.writerow([d["inventory_id"],d["filename"],d["extension"],d["relative_path"],d["repository"],d["file_size"],d["last_modified"],d["source_package"],d["object_type"],d["version"],d["checksum"],d["duplicate_candidate"],d["review_required"],d["lifecycle"],d["status"],d["notes"]])

def write_xlsx(records,workbook:Path):
    wb=load_workbook(workbook)
    ws=wb.active
    if ws.max_row>1:
        ws.delete_rows(2,ws.max_row-1)
    for r in records:
        d=r.to_dict()
        ws.append([d["inventory_id"],d["filename"],d["extension"],d["relative_path"],d["repository"],d["file_size"],d["last_modified"],d["source_package"],d["object_type"],d["version"],d["checksum"],d["duplicate_candidate"],d["review_required"],d["lifecycle"],d["status"],d["notes"]])
    wb.save(workbook)
